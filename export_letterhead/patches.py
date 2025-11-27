"""
Export Letterhead Patches

This module patches Frappe's export functions to add letterhead rows and font formatting
to all Excel and CSV exports. It intercepts export calls and prepends letterhead rows
with custom templates, then applies font settings to all rows.

Patched Functions:
- make_xlsx: Excel file generation (used by query reports, report views)
- build_xlsx_response: Excel response builder
- build_csv_response: CSV response builder
- get_csv_bytes: CSV bytes generator (used by query reports)
- query_report._export_query: Query report exports
- reportview._export_query: Report view exports

How It Works:
1. Patches are applied on module import and via boot_session hook
2. When export functions are called, letterhead rows are generated from template
3. Letterhead rows are prepended to export data
4. Font settings (name and size) are applied to ALL rows in Excel exports
5. Original functions are called with modified data

Example Flow:
    User exports report → _export_query_with_letterhead() called
    → Context built with doctype/report info
    → Letterhead rows generated from template
    → Data prepended with letterhead
    → make_xlsx() called with letterhead + data
    → Font applied to all rows
    → Excel file returned
"""

import frappe
from export_letterhead.utils import _get_settings, _build_context, _generate_letterhead_rows
from io import BytesIO


# Store original functions before patching (for restoration if needed)
_original_functions = {}


def _get_param_value(source, *keys, default=None):
    """
    Safely fetch parameter values from dict / frappe._dict / objects.
    """
    if not source:
        return default

    for key in keys:
        value = None
        if isinstance(source, dict):
            value = source.get(key)
        else:
            try:
                value = getattr(source, key)
            except AttributeError:
                value = None

        if isinstance(value, str):
            value = value.strip()

        if value not in (None, ""):
            return value

    return default


def apply_patches():
    """
    Apply all patches for letterhead export functionality.
    
    Patches Frappe's export functions to intercept Excel/CSV generation
    and add letterhead rows with font formatting. Only applies patches once
    to avoid duplicate patching.
    
    Patched Functions:
    - frappe.utils.xlsxutils.make_xlsx: Core Excel generation
    - frappe.utils.xlsxutils.build_xlsx_response: Excel HTTP response
    - frappe.utils.csvutils.build_csv_response: CSV HTTP response
    - frappe.desk.utils.get_csv_bytes: CSV bytes for reports
    - frappe.desk.query_report._export_query: Query report exports
    - frappe.desk.reportview._export_query: Report view exports
    """
    # Only apply patches once
    if hasattr(apply_patches, '_applied'):
        return
    apply_patches._applied = True
    
    # Patch make_xlsx
    from frappe.utils import xlsxutils
    if 'make_xlsx' not in _original_functions:
        _original_functions['make_xlsx'] = xlsxutils.make_xlsx
        xlsxutils.make_xlsx = _make_xlsx_with_letterhead
    
    # Patch build_xlsx_response
    if 'build_xlsx_response' not in _original_functions:
        _original_functions['build_xlsx_response'] = xlsxutils.build_xlsx_response
        xlsxutils.build_xlsx_response = _build_xlsx_response_with_letterhead
    
    # Patch build_csv_response
    from frappe.utils import csvutils
    if 'build_csv_response' not in _original_functions:
        _original_functions['build_csv_response'] = csvutils.build_csv_response
        csvutils.build_csv_response = _build_csv_response_with_letterhead
    
    # Patch get_csv_bytes (used by query reports)
    from frappe.desk import utils as desk_utils
    if 'get_csv_bytes' not in _original_functions:
        _original_functions['get_csv_bytes'] = desk_utils.get_csv_bytes
        desk_utils.get_csv_bytes = _get_csv_bytes_with_letterhead
    
    # Patch query_report._export_query to pass better context
    from frappe.desk import query_report
    if '_export_query' not in _original_functions:
        _original_functions['_export_query'] = query_report._export_query
        query_report._export_query = _export_query_with_letterhead
    
    # Patch reportview._export_query to pass better context
    from frappe.desk import reportview
    if '_export_query_reportview' not in _original_functions:
        _original_functions['_export_query_reportview'] = reportview._export_query
        reportview._export_query = _export_query_with_letterhead_reportview


def _make_xlsx_with_letterhead(data, sheet_name, wb=None, column_widths=None):
    """
    Wrapper for make_xlsx to add letterhead rows and apply font formatting.
    
    This function intercepts Excel file generation and:
    1. Generates letterhead rows from template
    2. Prepends letterhead rows to data
    3. Calls original make_xlsx function
    4. Applies font settings (name and size) to ALL rows
    
    Args:
        data: List of rows (each row is a list of cell values)
        sheet_name: Name of the Excel sheet
        wb: Optional existing workbook
        column_widths: Optional list of column widths
    
    Returns:
        BytesIO object containing Excel file with letterhead and font formatting
    
    Example:
        Input data: [["Header1", "Header2"], ["Value1", "Value2"]]
        Letterhead: [["Company Name"], ["Printed by: User"]]
        Output: Excel file with letterhead rows at top, all rows with custom font
    """
    settings = _get_settings()
    letterhead_rows_count = 0
    
    if settings and settings.get("enabled"):
        # Try to get context from frappe.local if set by higher-level functions
        export_context = getattr(frappe.local, 'export_letterhead_context', None)
        
        # Build context from available data
        if export_context:
            context = export_context.copy()
        else:
            context = _build_context((data, sheet_name), {"doctype": sheet_name, "report_name": sheet_name})
        
        # Generate letterhead rows (includes template rows + "Printed by" row if enabled)
        letterhead_rows = _generate_letterhead_rows(settings, context)
        
        if letterhead_rows:
            # Count ALL letterhead rows (template rows + "Printed by" row)
            letterhead_rows_count = len(letterhead_rows)
            # Prepend letterhead rows to data
            data = list(letterhead_rows) + list(data)
        
        # Clear context after use
        if hasattr(frappe.local, 'export_letterhead_context'):
            delattr(frappe.local, 'export_letterhead_context')
    
    # Call original function
    xlsx_file = _original_functions['make_xlsx'](data, sheet_name, wb, column_widths)
    
    # Apply font settings to ALL rows if enabled (letterhead + data rows)
    if settings and settings.get("enabled"):
        xlsx_file = _apply_font_to_all_rows(xlsx_file, settings, letterhead_rows_count)
    
    return xlsx_file


def _apply_font_to_all_rows(xlsx_file, settings, letterhead_rows_count):
    """
    Apply font settings (name and size) to ALL rows in Excel file.
    
    Loads the Excel file, applies configured font to every row (letterhead + data),
    and saves it back. Font is applied at both row and cell level for consistency.
    
    Args:
        xlsx_file: BytesIO object containing Excel file
        settings: Dictionary with font_name and font_size
        letterhead_rows_count: Number of letterhead rows (for reference, not used)
    
    Returns:
        BytesIO object with font formatting applied
    
    Font Settings:
        - font_name: Font family (e.g., "Arial", "Calibri", "Times New Roman")
        - font_size: Font size in points (1-409, default: 11)
    
    Example:
        settings = {"font_name": "Calibri", "font_size": 12}
        → All rows in Excel file will use Calibri 12pt font
    """
    try:
        from openpyxl import load_workbook
        from openpyxl.styles import Font
        
        # Reset file pointer
        xlsx_file.seek(0)
        
        # Load workbook (read-write mode to allow modifications)
        wb = load_workbook(xlsx_file, read_only=False)
        ws = wb.active
        
        # Get font settings and ensure they're the correct type
        font_name = settings.get("font_name", "Arial")
        if font_name:
            font_name = str(font_name).strip()
        if not font_name:
            font_name = "Arial"
        
        # Validate font name (remove any invalid characters)
        # Excel font names are typically alphanumeric with spaces and hyphens
        import re
        font_name = re.sub(r'[^\w\s\-]', '', font_name).strip()
        if not font_name:
            font_name = "Arial"
        
        # Ensure font_size is an integer and within valid range
        font_size = settings.get("font_size", 11)
        try:
            font_size = int(font_size) if font_size else 11
            # Excel font size range is typically 1-409
            if font_size < 1:
                font_size = 11
            elif font_size > 409:
                font_size = 409
        except (ValueError, TypeError):
            font_size = 11
        
        # Create font object once for efficiency
        export_font = Font(name=font_name, size=font_size)
        
        # Get the maximum row number in the worksheet
        max_row = ws.max_row
        
        # Apply font to ALL rows in the worksheet (letterhead + data)
        # Rows are 1-indexed in openpyxl
        for row_idx in range(1, max_row + 1):
            try:
                # Apply to row dimensions (affects entire row)
                row_dim = ws.row_dimensions[row_idx]
                row_dim.font = export_font
                
                # Also apply to each cell in the row to ensure it takes effect
                # This is important because cell-level font overrides row-level font
                for cell in ws[row_idx]:
                    if cell:
                        cell.font = export_font
            except Exception as row_error:
                # Log but continue with other rows
                frappe.logger("letterhead").debug(
                    f"Failed to apply font to row {row_idx}: {row_error}"
                )
        
        # Save to new BytesIO
        new_file = BytesIO()
        wb.save(new_file)
        new_file.seek(0)
        wb.close()  # Close workbook to free resources
        return new_file
    except Exception as e:
        frappe.logger("letterhead").error(
            f"Failed to apply font to all rows: {e}",
            exc_info=True
        )
        # Return original file if font application fails
        try:
            xlsx_file.seek(0)
        except Exception:
            pass
        return xlsx_file


def _build_xlsx_response_with_letterhead(data, filename):
    """Wrapper for build_xlsx_response to add letterhead rows"""
    settings = _get_settings()
    
    if settings and settings.get("enabled"):
        # Try to get context from frappe.local if set by higher-level functions
        export_context = getattr(frappe.local, 'export_letterhead_context', None)
        
        if export_context:
            context = export_context.copy()
        else:
            # Build context
            context = _build_context((data, filename), {"doctype": filename, "report_name": filename})
        
        # Generate letterhead rows
        letterhead_rows = _generate_letterhead_rows(settings, context)
        
        if letterhead_rows:
            # Prepend letterhead rows to data
            data = list(letterhead_rows) + list(data)
        
        # Clear context after use
        if hasattr(frappe.local, 'export_letterhead_context'):
            delattr(frappe.local, 'export_letterhead_context')
    
    # Call original function which uses make_xlsx
    # The font will be applied by make_xlsx wrapper
    return _original_functions['build_xlsx_response'](data, filename)


def _build_csv_response_with_letterhead(data, filename):
    """
    Wrapper for build_csv_response to add letterhead rows.
    
    Adds letterhead rows to CSV data before building HTTP response.
    Note: CSV files don't support font formatting, only content is added.
    
    Args:
        data: List of rows for CSV export
        filename: Name for the exported file
    
    Returns:
        HTTP response with CSV file containing letterhead rows
    """
    settings = _get_settings()
    
    if settings and settings.get("enabled"):
        # Try to get context from frappe.local if set by higher-level functions
        export_context = getattr(frappe.local, 'export_letterhead_context', None)
        
        if export_context:
            context = export_context.copy()
        else:
            # Build context
            context = _build_context((data, filename), {"doctype": filename, "report_name": filename})
        
        # Generate letterhead rows
        letterhead_rows = _generate_letterhead_rows(settings, context)
        
        if letterhead_rows:
            # Prepend letterhead rows to data
            data = list(letterhead_rows) + list(data)
        
        # Clear context after use
        if hasattr(frappe.local, 'export_letterhead_context'):
            delattr(frappe.local, 'export_letterhead_context')
    
    # Call original function
    return _original_functions['build_csv_response'](data, filename)


def _get_csv_bytes_with_letterhead(data, csv_params):
    """
    Wrapper for get_csv_bytes to add letterhead rows.
    
    Used by query reports for CSV export. Adds letterhead rows to data
    before converting to CSV bytes. Context is extracted from frappe.local
    if set by higher-level export functions.
    
    Args:
        data: List of rows for CSV export
        csv_params: Dictionary with CSV parameters (delimiter, quoting, etc.)
    
    Returns:
        Bytes object containing CSV data with letterhead rows
    """
    settings = _get_settings()
    
    if settings and settings.get("enabled"):
        # Try to get context from frappe.local if set by higher-level functions
        export_context = getattr(frappe.local, 'export_letterhead_context', None)
        
        if export_context:
            context = export_context.copy()
        else:
            # Build context - try to get doctype from data if possible
            doctype = "Export"
            if isinstance(csv_params, dict) and csv_params.get("doctype"):
                doctype = csv_params.get("doctype")
            
            context = _build_context((data,), {"doctype": doctype, "report_name": doctype})
        
        # Generate letterhead rows
        letterhead_rows = _generate_letterhead_rows(settings, context)
        
        if letterhead_rows:
            # Prepend letterhead rows to data
            data = list(letterhead_rows) + list(data)
        
        # Clear context after use
        if hasattr(frappe.local, 'export_letterhead_context'):
            delattr(frappe.local, 'export_letterhead_context')
    
    # Call original function
    return _original_functions['get_csv_bytes'](data, csv_params)


def _export_query_with_letterhead(form_params, csv_params, populate_response=True):
    """
    Wrapper for query_report._export_query to add letterhead context.
    
    Extracts report information and sets context in frappe.local for
    letterhead template rendering. This ensures report name and doctype
    are available in templates.
    
    Args:
        form_params: Form parameters containing report_name, filters, etc.
        csv_params: CSV export parameters
        populate_response: Whether to populate HTTP response
    
    Returns:
        Result from original _export_query function
    
    Context Variables Set:
        - doctype: Reference doctype from report (if available)
        - report_name: Name of the report being exported
    
    Example:
        Report "Sales Invoice Report" with ref_doctype "Sales Invoice"
        → Context: {"doctype": "Sales Invoice", "report_name": "Sales Invoice Report"}
        → Template can use: {{ report_name }} or {{ doctype }}
    """
    # Set context for letterhead generation
    try:
        report_name = _get_param_value(form_params, "report_name", "report")
        if not report_name:
            report_name = _get_param_value(getattr(frappe.local, "form_dict", None), "report_name", "report")
        if not report_name:
            report_name = "Query Report"

        # Try to get ref_doctype from report
        try:
            report_doc = frappe.get_doc("Report", report_name)
            ref_doctype = getattr(report_doc, "ref_doctype", None)
        except Exception:
            ref_doctype = None
        
        frappe.local.export_letterhead_context = _build_context(
            (),
            {
                "doctype": ref_doctype or report_name,
                "report_name": report_name,
            }
        )
    except Exception:
        pass
    
    try:
        # Call original function
        return _original_functions['_export_query'](form_params, csv_params, populate_response)
    finally:
        # Clean up context
        if hasattr(frappe.local, 'export_letterhead_context'):
            delattr(frappe.local, 'export_letterhead_context')


def _export_query_with_letterhead_reportview(form_params, csv_params, populate_response=True):
    """
    Wrapper for reportview._export_query to add letterhead context.
    
    Extracts doctype from form parameters and sets context for letterhead
    template rendering. Used for report view exports (list view exports).
    
    Args:
        form_params: Form parameters containing doctype, filters, fields, etc.
        csv_params: CSV export parameters
        populate_response: Whether to populate HTTP response
    
    Returns:
        Result from original reportview._export_query function
    
    Context Variables Set:
        - doctype: Document type being exported (e.g., "Sales Invoice", "Customer")
    
    Example:
        Exporting "Sales Invoice" list view
        → Context: {"doctype": "Sales Invoice"}
        → Template can use: {{ doctype }}
    """
    # Set context for letterhead generation
    try:
        doctype = _get_param_value(form_params, "doctype")
        report_label = _get_param_value(form_params, "report_name", "title", "report") or doctype
        if doctype:
            frappe.local.export_letterhead_context = _build_context(
                (),
                {"doctype": doctype, "report_name": report_label or doctype}
            )
    except Exception:
        pass
    
    try:
        # Call original function
        return _original_functions['_export_query_reportview'](form_params, csv_params, populate_response)
    finally:
        # Clean up context
        if hasattr(frappe.local, 'export_letterhead_context'):
            delattr(frappe.local, 'export_letterhead_context')


def boot_session(bootinfo):
    """
    Boot session hook to ensure patches are applied during web boot.
    
    Called by Frappe during application startup to apply patches
    even if module wasn't imported yet. This ensures letterhead
    functionality works in all scenarios.
    
    Args:
        bootinfo: Boot information dictionary
    
    Returns:
        Modified bootinfo dictionary
    """
    apply_patches()
    return bootinfo
